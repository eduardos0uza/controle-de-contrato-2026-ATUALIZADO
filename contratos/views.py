import json
import csv
from datetime import timedelta

from django.http import HttpResponse, JsonResponse
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import IntegrityError, transaction
from django.db.models import Count, Sum, Q, F, ExpressionWrapper, DecimalField, Value
from django.db.models.functions import Coalesce, TruncMonth
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from django.core.exceptions import ObjectDoesNotExist
from .db_utils import is_sequence_integrity_error, reset_contract_sequences
from .forms import ContractForm, ContractItemFormSet
from .models import AlertStatus, Contract, ContractHistory, ContractStatus, ContractItem
from .services import atualizar_alertas_diarios
from admin_panel.models import UserProfile
from admin_panel.utils import log_audit

User = get_user_model()

def get_user_contracts(user):
    """Retorna os contratos que o usuário tem permissão para ver."""
    # Superusuários e membros do grupo 'administrador' veem tudo
    if user.is_superuser or user.groups.filter(name='administrador').exists():
        return Contract.objects.all()
    
    try:
        profile = user.profile
        if profile and profile.secretaria:
            return Contract.objects.filter(secretaria=profile.secretaria)
    except (ObjectDoesNotExist, AttributeError):
        pass
    
    return Contract.objects.none()

def _tem_roles(user, roles):
    return user.is_superuser or user.groups.filter(name__in=roles).exists()


def _roles_required(roles):
    def check(user):
        return _tem_roles(user, roles)

    return user_passes_test(check)


def _salvar_contrato_com_itens(form, formset, user=None):
    try:
        with transaction.atomic():
            contrato = form.save(commit=False)
            if user:
                contrato.criado_por = user
                # Se o usuário não for superuser, força a secretaria dele no contrato
                if not user.is_superuser and hasattr(user, 'profile') and user.profile.secretaria:
                    contrato.secretaria = user.profile.secretaria
            contrato.save()
            formset.instance = contrato
            formset.save()
            return contrato
    except IntegrityError as exc:
        if not is_sequence_integrity_error(exc):
            raise

    reset_contract_sequences()
    form.instance.pk = None
    form.instance.id = None
    with transaction.atomic():
        contrato = form.save(commit=False)
        if user:
            contrato.criado_por = user
            if not user.is_superuser and hasattr(user, 'profile') and user.profile.secretaria:
                contrato.secretaria = user.profile.secretaria
        contrato.save()
        formset.instance = contrato
        formset.save()
        return contrato


def landing_page(request):
    if request.user.is_authenticated:
        return redirect('contratos:dashboard')
    return render(request, 'home.html')

@login_required
def dashboard(request):
    try:
        # Atualizar status dos contratos diariamente ao acessar o dashboard
        # Isso garante que as informações estejam sempre corretas mesmo se o servidor for desligado
        try:
            atualizar_alertas_diarios()
        except Exception as e_alert:
            print(f"Erro ao atualizar alertas: {e_alert}")

        contratos = get_user_contracts(request.user)
        hoje = timezone.localdate()
        
        # Calculate counts in a more efficient way (fewer queries)
        try:
            metrics = contratos.aggregate(
                total_contratos=Count('id'),
                total_alertas=Count('id', filter=~Q(alerta='normal')),
                vencidos=Count('id', filter=Q(data_vencimento__lt=hoje)),
                vencendo_30=Count('id', filter=Q(data_vencimento__gte=hoje, data_vencimento__lte=hoje + timedelta(days=30)))
            )
            
            total_contratos = metrics.get('total_contratos', 0)
            total_alertas = metrics.get('total_alertas', 0)
            vencidos = metrics.get('vencidos', 0)
            vencendo_30 = metrics.get('vencendo_30', 0)
        except Exception as e_metrics:
            print(f"Erro ao calcular métricas: {e_metrics}")
            total_contratos = total_alertas = vencidos = vencendo_30 = 0

        # These still need separate queries because of grouping
        contagem_alertas = list(contratos.values('alerta').annotate(total=Count('id')))
        contagem_status = list(contratos.values('status').annotate(total=Count('id')))
        
        proximos_vencimento = contratos.filter(dias_restantes__lte=90).order_by('dias_restantes')[:10]
        
        # Prepare chart data with labels
        status_map = dict(ContractStatus.choices)
        
        # Map internal values to display labels consistent with template
        alert_display_map = {
            'normal': 'Normal',
            'amarelo': 'Alerta (90d)',
            'laranja': 'Atenção (30d)',
            'vermelho': 'Crítico (7d)',
            'vencido': 'Vencido'
        }
        
        # Descrições para tooltips/legendas
        alert_description_map = {
            'normal': 'Prazo seguro (mais de 90 dias)',
            'amarelo': 'Vence em menos de 3 meses',
            'laranja': 'Vence no próximo mês',
            'vermelho': 'Vence nesta semana',
            'vencido': 'Prazo expirado'
        }
        
        # Ordem lógica: Normal -> Alerta -> Atenção -> Crítico -> Vencido
        ordem_alertas = ['normal', 'amarelo', 'laranja', 'vermelho', 'vencido']
        dist_alertas_dict = {item['alerta']: item['total'] for item in contagem_alertas}
        
        alertas_chart = [
            {'label': alert_display_map.get(key, key), 'value': dist_alertas_dict.get(key, 0), 'type': key}
            for key in ordem_alertas
            if dist_alertas_dict.get(key, 0) > 0
        ]
        
        status_chart = [
            {'label': status_map.get(key, key), 'value': next((x['total'] for x in contagem_status if x['status'] == key), 0), 'type': key}
            for key, _ in ContractStatus.choices
        ]

        context = {
            'contagem_alertas': contagem_alertas,
            'contagem_status': contagem_status,
            'total_alertas': total_alertas,
            'total_contratos': total_contratos,
            'vencidos': vencidos,
            'vencendo_30': vencendo_30,
            'proximos_vencimento': proximos_vencimento,
            'alertas_chart': alertas_chart,
            'status_chart': status_chart,
            'alert_display_map': alert_display_map,
            'alert_description_map': alert_description_map,
        }

        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('api') == 'true':
            return JsonResponse({
                'vencendo_30': vencendo_30,
                'total_alertas': total_alertas,
                'vencidos': vencidos,
                'alertas_chart': alertas_chart,
                'status_chart': status_chart,
            })

        return render(request, 'contratos/dashboard.html', context)
    except Exception as e:
        print(f"ERRO CRÍTICO DASHBOARD: {e}")
        import traceback
        print(traceback.format_exc())
        return render(request, '500.html', {'error': str(e)}, status=500)


@login_required
def lista_contratos(request):
    try:
        contratos = get_user_contracts(request.user).annotate(
            valor_contrato=Coalesce(Sum('itens__valor_total'), Value(0), output_field=DecimalField())
        )
        
        # Filtros
        search_term = request.GET.get('search')
        inicio = request.GET.get('inicio')
        fim = request.GET.get('fim')
        status = request.GET.get('status')
        responsavel = request.GET.get('responsavel')
        alerta = request.GET.get('alerta')
        
        if search_term:
            contratos = contratos.filter(
                Q(numero_contrato__icontains=search_term) |
                Q(objeto__icontains=search_term) |
                Q(responsavel__icontains=search_term)
            )
        
        if inicio:
            contratos = contratos.filter(data_vencimento__gte=inicio)
        if fim:
            contratos = contratos.filter(data_vencimento__lte=fim)
        if status:
            contratos = contratos.filter(status=status)
        if responsavel:
            contratos = contratos.filter(responsavel__icontains=responsavel)
        if alerta:
            contratos = contratos.filter(alerta=alerta)

        # Ordenação
        sort_field = request.GET.get('sort', 'dias_restantes')
        sort_direction = request.GET.get('direction', 'asc')
        
        valid_sort_fields = {
            'numero_contrato': 'numero_contrato',
            'data_vencimento': 'data_vencimento',
            'valor': 'valor_contrato',
            'status': 'status',
            'responsavel': 'responsavel',
            'dias_restantes': 'dias_restantes',
        }

        if sort_field in valid_sort_fields:
            db_field = valid_sort_fields[sort_field]
            if sort_direction == 'desc':
                db_field = f'-{db_field}'
            contratos = contratos.order_by(db_field, '-id')
        else:
            # Default sort (urgência)
            contratos = contratos.order_by('dias_restantes', 'numero_contrato')

        # KPIs dos resultados filtrados
        total_exibidos = contratos.count()
        valor_total_exibido = ContractItem.objects.filter(contrato__in=contratos).aggregate(total=Sum('valor_total'))['total'] or 0
        hoje = timezone.localdate()
        vencendo_30_exibido = contratos.filter(data_vencimento__gte=hoje, data_vencimento__lte=hoje + timedelta(days=30)).count()
        vencidos_exibido = contratos.filter(data_vencimento__lt=hoje).count()

        # Paginação
        per_page = request.GET.get('per_page', 25)
        try:
            per_page = int(per_page)
            if per_page not in [10, 25, 50, 100]:
                per_page = 25
        except ValueError:
            per_page = 25

        paginator = Paginator(contratos, per_page)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Check for partial rendering for real-time updates
        if request.GET.get('partial'):
            return render(request, 'contratos/partials/_lista_contratos_tbody.html', {'contratos': page_obj})

        context = {
            'contratos': page_obj, # Agora passamos o objeto paginado
            'status_choices': ContractStatus.choices,
            'alerta_choices': AlertStatus.choices,
            'responsaveis': User.objects.order_by('username').values('id', 'username'),
            'filtros': {
                'inicio': inicio, 
                'fim': fim, 
                'status': status, 
                'responsavel': responsavel, 
                'alerta': alerta, 
                'search': search_term, 
                'per_page': per_page,
                'sort': sort_field,
                'direction': sort_direction
            },
            'kpis': {
                'total': total_exibidos,
                'valor': valor_total_exibido,
                'vencendo_30': vencendo_30_exibido,
                'vencidos': vencidos_exibido,
            }
        }
        return render(request, 'contratos/lista_contratos.html', context)
    except Exception as e:
        print(f"ERRO lista_contratos: {e}")
        return render(request, '500.html', {'error': str(e)}, status=500)


@login_required
@_roles_required(['administrador', 'gestor'])
def criar_contrato(request):
    if request.method == 'POST':
        form = ContractForm(request.POST, request.FILES, request_user=request.user)
        formset = ContractItemFormSet(request.POST, request.FILES)
        
        if form.is_valid() and formset.is_valid():
            # Atribuir secretaria do usuário ao contrato se ele tiver uma
            try:
                if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.secretaria:
                    form.instance.secretaria = request.user.profile.secretaria
            except UserProfile.DoesNotExist:
                pass
                
            _salvar_contrato_com_itens(form, formset, user=request.user)
            return redirect('contratos:lista')
    else:
        form = ContractForm(request_user=request.user)
        formset = ContractItemFormSet()
    
    return render(request, 'contratos/contrato_form.html', {
        'form': form,
        'formset': formset,
        'titulo': 'Novo Contrato'
    })

@login_required
@_roles_required(['administrador', 'gestor'])
def editar_contrato(request, contrato_id):
    contratos_permitidos = get_user_contracts(request.user)
    contrato = get_object_or_404(contratos_permitidos, pk=contrato_id)
    
    if request.method == 'POST':
        form = ContractForm(request.POST, request.FILES, instance=contrato, request_user=request.user)
        formset = ContractItemFormSet(request.POST, request.FILES, instance=contrato)
        
        if form.is_valid() and formset.is_valid():
            # Reforçar secretaria para não-superusers
            if not request.user.is_superuser:
                try:
                    if hasattr(request.user, 'profile') and request.user.profile.secretaria:
                        form.instance.secretaria = request.user.profile.secretaria
                except UserProfile.DoesNotExist:
                    pass
            
            _salvar_contrato_com_itens(form, formset, user=request.user)
            return redirect('contratos:detalhe', contrato_id=contrato.id)
    else:
        form = ContractForm(instance=contrato, request_user=request.user)
        formset = ContractItemFormSet(instance=contrato)
    
    return render(request, 'contratos/contrato_form.html', {
        'form': form,
        'formset': formset,
        'titulo': f'Editar Contrato: {contrato.numero_contrato}'
    })


@login_required
@_roles_required(['administrador', 'gestor'])
def duplicar_contrato(request, contrato_id):
    contratos_permitidos = get_user_contracts(request.user)
    contrato_origem = get_object_or_404(contratos_permitidos, pk=contrato_id)
    
    if request.method == 'POST':
        form = ContractForm(request.POST, request.FILES, request_user=request.user)
        formset = ContractItemFormSet(request.POST, request.FILES)
        if form.is_valid() and formset.is_valid():
            # Atribuir secretaria do usuário ao novo contrato (mesma lógica de criar_contrato)
            try:
                if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.secretaria:
                    form.instance.secretaria = request.user.profile.secretaria
                else:
                    # Se for superuser duplicando, talvez queira manter a secretaria original
                    form.instance.secretaria = contrato_origem.secretaria
            except UserProfile.DoesNotExist:
                pass
                
            _salvar_contrato_com_itens(form, formset, user=request.user)
            return redirect('contratos:lista')
    else:
        # Pre-fill form with original data but tweaked for duplication
        # Criamos uma nova instância baseada na original para preencher o formulário
        dados_origem = Contract.objects.get(pk=contrato_id)
        dados_origem.pk = None
        dados_origem.id = None
        dados_origem.numero_contrato = f"{dados_origem.numero_contrato} (Cópia)"
        
        form = ContractForm(instance=dados_origem, request_user=request.user)
        
        # Prepare initial data for formset based on original items
        # We use initial because we are creating NEW items, not editing old ones
        original_items = ContractItem.objects.filter(contrato_id=contrato_id).values(
            'descricao', 'unidade', 'quantidade', 'valor_unitario', 'valor_total'
        )
        formset = ContractItemFormSet(queryset=ContractItem.objects.none(), initial=list(original_items))
    
    return render(request, 'contratos/contrato_form.html', {
        'form': form,
        'formset': formset,
        'titulo': 'Duplicar Contrato'
    })


@login_required
def renovar_contrato(request, contrato_id):
    contratos_permitidos = get_user_contracts(request.user)
    contrato = get_object_or_404(contratos_permitidos, pk=contrato_id)
    
    if request.method == 'POST':
        from django.contrib import messages
        from django.db import transaction
        from contratos.forms import ContractRenewalForm
        from contratos.models import ContractItem
        
        form = ContractRenewalForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. Obter os dados do formulário
                    renovacao = form.save(commit=False)
                    renovacao.contrato = contrato
                    renovacao.responsavel = request.user
                    
                    # 2. Registrar valores anteriores
                    renovacao.vencimento_anterior = contrato.data_vencimento
                    renovacao.valor_anterior = contrato.valor
                    
                    # 3. Salvar o registro de renovação
                    renovacao.save()
                    
                    # 4. Atualizar o contrato real com os novos dados
                    contrato.data_vencimento = renovacao.vencimento_novo
                    
                    # Adicionar item orçamentário se houver diferença
                    valor_diferenca = renovacao.valor_novo - renovacao.valor_anterior
                    if valor_diferenca != 0:
                        ContractItem.objects.create(
                            contrato=contrato,
                            tipo_item='SERVICO',
                            descricao=f"Aditivo/Renovação — {renovacao.get_tipo_renovacao_display()}",
                            quantidade=1,
                            valor_unitario=valor_diferenca,
                            valor_total=valor_diferenca
                        )
                    
                    # Atualiza a preferência de renovação automática
                    contrato.renovacao_automatica = 'renovacao_automatica' in request.POST
                    contrato.save()
                    
                    messages.success(request, f"Contrato {contrato.numero_contrato} renovado com sucesso!")
                    return redirect('contratos:lista')
            except Exception as e:
                import traceback
                print(f"CRITICAL ERROR RENOVACAO_CONTRATO: {str(e)}")
                print(traceback.format_exc())
                messages.error(request, f"Erro interno ao processar renovação: {str(e)}")
                return redirect('contratos:lista')
        else:
            # Captura erros específicos do form para mostrar ao usuário
            error_msgs = []
            for field, errors in form.errors.items():
                field_name = form.fields[field].label if field in form.fields else field
                error_msgs.append(f"{field_name}: {', '.join(errors)}")
            
            error_text = " Erro nos campos: " + " | ".join(error_msgs)
            print(f"FORM INVALID RENOVACAO_CONTRATO: {error_text}")
            messages.error(request, f"Não foi possível renovar o contrato.{error_text}")
            return redirect('contratos:lista')
            
    # GET request para dados do modal (JSON)
    return JsonResponse({
        'id': contrato.id,
        'numero': contrato.numero_contrato,
        'empresa': contrato.empresa,
        'valor_atual': float(contrato.valor),
        'vencimento_atual': contrato.data_vencimento.strftime('%Y-%m-%d'),
        'dias_restantes': contrato.dias_restantes,
        'status': contrato.get_status_display(),
    })


@login_required
@_roles_required(['administrador', 'gestor'])
def excluir_contrato(request, contrato_id):
    try:
        contratos_permitidos = get_user_contracts(request.user)
        contrato = get_object_or_404(contratos_permitidos, pk=contrato_id)
        if request.method == 'POST':
            numero_contrato = contrato.numero_contrato
            contrato_id_str = str(contrato.id)
            contrato.delete()
            
            # Log de Auditoria
            log_audit(
                request, 
                'DELETE', 
                'Contract', 
                contrato_id_str, 
                f"Contrato {numero_contrato} foi excluído por {request.user.username}"
            )
            
            messages.success(request, f"Contrato {numero_contrato} excluído com sucesso.")
            return redirect('contratos:lista')
        return redirect('contratos:detalhe', contrato_id=contrato.id)
    except Exception as e:
        print(f"ERRO excluir_contrato: {e}")
        messages.error(request, "Erro ao excluir contrato.")
        return redirect('contratos:lista')


@login_required
def detalhe_contrato(request, contrato_id):
    contrato = get_object_or_404(get_user_contracts(request.user).prefetch_related('itens'), pk=contrato_id)
    
    # Cálculo dinâmico de progresso para o template
    if contrato.data_inicio and contrato.data_vencimento:
        total = (contrato.data_vencimento - contrato.data_inicio).days
        if total > 0:
            passados = (timezone.localdate() - contrato.data_inicio).days
            contrato.progresso = min(max(int((passados / total) * 100), 0), 100)
        else:
            contrato.progresso = 100
    else:
        contrato.progresso = 0
        
    historico = ContractHistory.objects.select_related('alterado_por').filter(contrato=contrato).order_by('-alterado_em')[:20]
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('partial') == 'true':
        return render(request, 'contratos/partials/_detalhe_contrato_content.html', {'contrato': contrato, 'historico': historico})
        
    return render(request, 'contratos/detalhe_contrato.html', {'contrato': contrato, 'historico': historico})


def privacidade(request):
    return render(request, 'contratos/privacidade.html')


def termos_uso(request):
    return render(request, 'contratos/termos_uso.html')


@login_required
def relatorios(request):
    hoje = timezone.localdate()
    
    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status_filtro = request.GET.get('status')
    
    todos_contratos = get_user_contracts(request.user)
    
    if data_inicio:
        todos_contratos = todos_contratos.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        todos_contratos = todos_contratos.filter(data_vencimento__lte=data_fim)
    if status_filtro:
        todos_contratos = todos_contratos.filter(status=status_filtro)
    contratos_com_valor = todos_contratos.annotate(
        valor_contrato=Coalesce(Sum('itens__valor_total'), Value(0), output_field=DecimalField())
    )
        
    # Exportação CSV
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="relatorio_contratos_{hoje}.csv"'
        
        # Mapeamento de colunas disponíveis
        # Chave: valor do checkbox, Valor: (Título do Header, Função para extrair dado)
        column_map = {
            'numero_contrato': ('Número', lambda c: c.numero_contrato),
            'numero_protocolo': ('Protocolo', lambda c: c.numero_protocolo or '-'),
            'processo_administrativo': ('Processo Adm.', lambda c: c.processo_administrativo or '-'),
            'numero_pregao': ('Pregão', lambda c: c.numero_pregao or '-'),
            'secretario': ('Secretário(a)', lambda c: c.secretario or '-'),
            'numero_nota_empenho': ('Empenho', lambda c: c.numero_nota_empenho or '-'),
            'numero_ficha': ('Ficha', lambda c: c.numero_ficha or '-'),
            'objeto': ('Descrição', lambda c: c.objeto),
            'unidade': ('Unidade', lambda c: c.unidade or '-'),
            'quantidade': ('Qtd.', lambda c: str(c.quantidade)),
            'valor_unitario': ('Valor Unit.', lambda c: f"{c.valor_unitario:.2f}".replace('.', ',')),
            'valor': ('Valor Total', lambda c: f"{c.valor_contrato:.2f}".replace('.', ',')),
            'data_inicio': ('Início', lambda c: c.data_inicio.strftime('%d/%m/%Y') if c.data_inicio else '-'),
            'vigencia': ('Vigência (meses)', lambda c: str(c.vigencia)),
            'data_vencimento': ('Vencimento', lambda c: c.data_vencimento.strftime('%d/%m/%Y') if c.data_vencimento else '-'),
            'status': ('Status', lambda c: c.get_status_display()),
            'responsavel': ('Responsável', lambda c: c.responsavel if c.responsavel else '-'),
            'alerta': ('Alerta', lambda c: c.get_alerta_display()),
            'dias_restantes': ('Dias Restantes', lambda c: str(c.dias_restantes)),
        }

        # Colunas padrão se nenhuma for selecionada
        default_columns = ['numero_contrato', 'objeto', 'responsavel', 'valor', 'status', 'data_vencimento', 'alerta']
        
        # Obter colunas selecionadas da query string
        selected_keys = request.GET.getlist('columns')
        
        # Se não vier nada (exportação simples), usa padrão
        if not selected_keys:
            selected_keys = default_columns
            
        # Filtrar chaves válidas para evitar erros se usuário passar lixo na URL
        valid_keys = [k for k in selected_keys if k in column_map]
        
        # Se após filtro não sobrar nada, volta pro padrão
        if not valid_keys:
            valid_keys = default_columns

        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        
        # Escrever Header
        headers = [column_map[k][0] for k in valid_keys]
        writer.writerow(headers)
        
        # Escrever Linhas
        for c in contratos_com_valor:
            row = []
            for k in valid_keys:
                extract_func = column_map[k][1]
                row.append(extract_func(c))
            writer.writerow(row)
            
        return response
    
    # KPIs e Gráficos
    total_contratos = todos_contratos.count()
    valor_total = ContractItem.objects.filter(contrato__in=todos_contratos).aggregate(total=Sum('valor_total'))['total'] or 0
    contratos_ativos = todos_contratos.filter(status='ativo').count()

    # Novos KPIs solicitados
    vencendo_30d = todos_contratos.filter(data_vencimento__gte=hoje, data_vencimento__lte=hoje + timedelta(days=30)).count()
    vencidos = todos_contratos.filter(data_vencimento__lt=hoje).count()
    
    vencendo_30d_qs = todos_contratos.filter(data_vencimento__gte=hoje, data_vencimento__lte=hoje + timedelta(days=30))
    valor_vencendo_30d = ContractItem.objects.filter(contrato__in=vencendo_30d_qs).aggregate(total=Sum('valor_total'))['total'] or 0
    
    # Distribuição por Status
    dist_status = list(todos_contratos.values('status').annotate(total=Count('id')))
    
    # Distribuição por Responsável
    dist_responsavel = list(todos_contratos.values('responsavel').annotate(total=Count('id')).order_by('-total'))
    
    # Distribuição por Alerta
    dist_alertas = list(todos_contratos.values('alerta').annotate(total=Count('id')))

    # Timeline (Vencimentos por Mês)
    # Se não houver filtro de data, foca nos próximos 12 meses para o gráfico ser útil
    timeline_qs = todos_contratos
    if not data_inicio and not data_fim:
        timeline_qs = timeline_qs.filter(data_vencimento__gte=hoje, data_vencimento__lte=hoje + timedelta(days=365))
        
    dist_vencimentos = list(timeline_qs.annotate(mes=TruncMonth('data_vencimento'))
                           .values('mes')
                           .annotate(total=Count('id', distinct=True), valor_total=Sum('itens__valor_total'))
                           .order_by('mes'))
    
    # Preparar dados para gráficos (JSON)
    status_map = dict(ContractStatus.choices)
    status_data = [
        {
            'label': status_map.get(item['status'], item['status']), 
            'value': item['total'],
            'status': item['status']
        }
        for item in dist_status
    ]

    # Map internal values to display labels consistent with template
    alert_display_map = {
        'normal': 'Normal',
        'amarelo': 'Alerta (90d)',    # Was Alerta
        'laranja': 'Atenção (30d)',   # Was Atenção
        'vermelho': 'Crítico (7d)',   # Was Crítico
        'vencido': 'Vencido'
    }
    
    # Descrições para tooltips/legendas
    alert_description_map = {
        'normal': 'Prazo seguro (mais de 90 dias)',
        'amarelo': 'Vence em menos de 3 meses',
        'laranja': 'Vence no próximo mês',
        'vermelho': 'Vence nesta semana',
        'vencido': 'Prazo expirado'
    }
    
    # Ordem lógica: Normal -> Alerta -> Atenção -> Crítico -> Vencido
    ordem_alertas = ['normal', 'amarelo', 'laranja', 'vermelho', 'vencido']
    dist_alertas_dict = {item['alerta']: item['total'] for item in dist_alertas}
    
    alertas_data = [
        {'label': alert_display_map.get(key, key), 'value': dist_alertas_dict.get(key, 0), 'type': key}
        for key in ordem_alertas
        if dist_alertas_dict.get(key, 0) > 0
    ]

    responsavel_data = [
        {'label': item['responsavel'] or 'Sem Responsável', 'value': item['total']}
        for item in dist_responsavel
    ]

    MESES_PT = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    vencimentos_data = []
    for item in dist_vencimentos:
        mes = item['mes']
        if not mes:
            continue
        label = f"{MESES_PT[mes.month-1]}/{mes.year}"
        vencimentos_data.append({'mes': label, 'total': item['total'], 'valor': float(item['valor_total'] or 0)})

    # Lista Detalhada (Ordenada por vencimento)
    contratos_lista = contratos_com_valor.order_by('data_vencimento')
    
    # Calcular dias restantes e dias vencido no backend
    for c in contratos_lista:
        if c.data_vencimento:
            dias = (c.data_vencimento - hoje).days
            c.dias_restantes = dias
            # c.dias_vencido property handles the negative logic automatically
            # Property logic: if dias_restantes < 0: return abs(dias_restantes)
        else:
            c.dias_restantes = None
            # c.dias_vencido defaults to 0 via property logic if dias_restantes is 0 (default)
    
    # Configuração de Alertas (para legenda dinâmica)
    alertas_config = [
        {'type': 'normal', 'label': 'Regular', 'desc': 'Mais de 90 dias', 'color': 'var(--hiden-success)'},
        {'type': 'amarelo', 'label': 'Atenção', 'desc': 'Até 90 dias', 'color': '#ffc107'},
        {'type': 'laranja', 'label': 'Urgente', 'desc': 'Até 30 dias', 'color': 'var(--hiden-warning)'},
        {'type': 'vermelho', 'label': 'Crítico', 'desc': 'Até 7 dias', 'color': 'var(--hiden-danger)'},
        {'type': 'vencido', 'label': 'Expirado', 'desc': 'Já vencido', 'color': '#000000'},
    ]

    context = {
        'kpi_total': total_contratos,
        'kpi_valor': valor_total,
        'kpi_ativos': contratos_ativos,
        'kpi_vencendo_30d': vencendo_30d,
        'kpi_vencidos': vencidos,
        'kpi_valor_vencendo_30d': valor_vencendo_30d,
        'dist_status': dist_status,
        'dist_responsavel': dist_responsavel,
        'dist_alertas': dist_alertas,
        'dist_vencimentos': dist_vencimentos,
        # Dados para gráficos (para json_script)
        'status_data': status_data,
        'alertas_data': alertas_data,
        'responsavel_data': responsavel_data,
        'vencimentos_data': vencimentos_data,
        'alertas_config': alertas_config,
        'contratos': contratos_lista,
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'status': status_filtro
        },
        'status_choices': ContractStatus.choices,
    }
    return render(request, 'contratos/relatorios.html', context)




# =============================================================================
# MÓDULO — CALCULADORA GERENCIAL DE DISTRIBUIÇÃO MENSAL
# =============================================================================

from decimal import Decimal, ROUND_HALF_UP
from dateutil.relativedelta import relativedelta
from datetime import date
from .models import MESES_PT, MESES_PT_ABREV


def _format_currency_br(val):
    if val is None:
        return 'R$ 0,00'
    try:
        formatted = f"{float(val):.2f}"
        parts = formatted.split('.')
        integer_part = parts[0]
        decimal_part = parts[1]
        
        # Add thousands separator
        reversed_integer = integer_part[::-1]
        chunks = [reversed_integer[i:i+3] for i in range(0, len(reversed_integer), 3)]
        integer_with_sep = ".".join(chunks)[::-1]
        
        return f"R$ {integer_with_sep},{decimal_part}"
    except Exception:
        return 'R$ 0,00'


@login_required
def controle_mensal_calculo(request):
    """
    Visualização simplificada da distribuição gerencial mensal de todos os contratos cadastrados.
    Busca todos os contratos visíveis, lê o valor global (soma dos itens),
    permite selecionar mês/ano inicial e final, calcula a quantidade de meses do período
    e exibe uma tabela tipo planilha gerencial com colunas para cada mês.
    """
    # 1. Diagnóstico e logs temporários no backend
    import logging
    logger = logging.getLogger('django')
    
    total_contratos_db = Contract.objects.count()
    contratos_qs = get_user_contracts(request.user).prefetch_related('itens')
    contratos_ids_user = set(contratos_qs.values_list('id', flat=True))
    
    logger.info(f"[DIAGNOSTICO] Total de contratos cadastrados no Banco de Dados: {total_contratos_db}")
    
    ignored_details = []
    # Busca todos os contratos no banco de dados para listar os que este usuário não tem permissão para visualizar
    all_contracts = Contract.objects.all()
    for c in all_contracts:
        if c.id not in contratos_ids_user:
            reason = f"Ignorado por restrição de permissão: Contrato pertence à secretaria '{c.secretaria.nome}'" if c.secretaria else "Ignorado por restrição de permissão: Contrato sem secretaria atribuída (acesso restrito)"
            ignored_details.append((c.numero_contrato, reason))
            logger.info(f"[DIAGNOSTICO] Contrato {c.numero_contrato} ignorado. Motivo: {reason}")
            
    logger.info(f"[DIAGNOSTICO] Total de contratos considerados (acessíveis ao usuário {request.user.username}): {len(contratos_ids_user)}")
    logger.info(f"[DIAGNOSTICO] Total ignorado para este usuário: {len(ignored_details)}")

    # 2. Obter período selecionado ou definir padrão (junho a dezembro de 2026)
    hoje = timezone.localdate()
    
    # Valores padrão de acordo com o exemplo solicitado
    mes_inicio_padrao = 6
    ano_inicio_padrao = 2026
    mes_fim_padrao = 12
    ano_fim_padrao = 2026
    
    try:
        mes_inicio = int(request.GET.get('mes_inicio') or mes_inicio_padrao)
        ano_inicio = int(request.GET.get('ano_inicio') or ano_inicio_padrao)
        mes_fim    = int(request.GET.get('mes_fim') or mes_fim_padrao)
        ano_fim    = int(request.GET.get('ano_fim') or ano_fim_padrao)
    except (ValueError, TypeError):
        mes_inicio = mes_inicio_padrao
        ano_inicio = ano_inicio_padrao
        mes_fim = mes_fim_padrao
        ano_fim = ano_fim_padrao

    try:
        start_date = date(ano_inicio, mes_inicio, 1)
        end_date   = date(ano_fim, mes_fim, 1)
    except Exception:
        start_date = date(2026, 6, 1)
        end_date   = date(2026, 12, 1)
        mes_inicio, ano_inicio = 6, 2026
        mes_fim, ano_fim = 12, 2026

    if start_date > end_date:
        start_date, end_date = end_date, start_date
        mes_inicio, mes_fim = mes_fim, mes_inicio
        ano_inicio, ano_fim = ano_fim, ano_inicio

    # 3. Gerar lista de meses/colunas do período selecionado
    colunas_meses = []
    temp_date = start_date
    while temp_date <= end_date:
        label_compacto = f"{MESES_PT_ABREV[temp_date.month].upper()}/{temp_date.year}"
        label_very_compact = f"{MESES_PT_ABREV[temp_date.month].upper()}/{str(temp_date.year)[2:]}"
        colunas_meses.append({
            'mes': temp_date.month,
            'ano': temp_date.year,
            'label': f"{MESES_PT[temp_date.month]}/{temp_date.year}",
            'label_compacto': label_compacto,
            'label_very_compact': label_very_compact,
            'mes_nome': MESES_PT[temp_date.month],
            'is_current': (temp_date.month == hoje.month and temp_date.year == hoje.year)
        })
        temp_date += relativedelta(months=1)

    total_meses_periodo = len(colunas_meses)
    if total_meses_periodo <= 0:
        total_meses_periodo = 1

    # 4. Processar contratos
    dados_tabela = []
    totais_por_mes = {f"{c['mes']}_{c['ano']}": Decimal('0.00') for c in colunas_meses}
    total_global_acumulado = Decimal('0.00')

    for contrato in contratos_qs:
        valor_global = Decimal(str(contrato.valor or 0))
        # Não excluir contrato mesmo se valor_global <= 0, fornecedor/secretaria/vigência em branco, ou fora do período.

        # Calcular valor mensal = valor global / quantidade de meses do período
        valor_mensal = (valor_global / total_meses_periodo).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        tem_vigencia = bool(contrato.data_inicio and contrato.data_vencimento)
        
        # Obter os valores distribuídos para cada coluna de mês
        valores_distribuidos = {}
        is_active_dict = {}
        for col in colunas_meses:
            col_key = f"{col['mes']}_{col['ano']}"
            col_date = date(col['ano'], col['mes'], 1)
            
            # Verificar se o mês da coluna está dentro da vigência do contrato
            is_active = False
            if tem_vigencia:
                contrato_start = contrato.data_inicio.replace(day=1)
                contrato_end = contrato.data_vencimento.replace(day=1)
                is_active = (contrato_start <= col_date <= contrato_end)
            
            is_active_dict[col_key] = is_active
            if is_active:
                valores_distribuidos[col_key] = valor_mensal
                totais_por_mes[col_key] += valor_mensal
            else:
                valores_distribuidos[col_key] = Decimal('0.00')

        total_global_acumulado += valor_global
        
        dados_tabela.append({
            'contrato': contrato,
            'numero': contrato.numero_contrato,
            'fornecedor': contrato.empresa or 'Não informado',
            'secretaria': contrato.secretaria.nome if contrato.secretaria else 'Não informado',
            'valor_global_raw': valor_global,
            'valor_global': _format_currency_br(valor_global),
            'tem_vigencia': tem_vigencia,
            'valores_meses': [
                {
                    'raw': valores_distribuidos[f"{c['mes']}_{c['ano']}"],
                    'formatted': _format_currency_br(valores_distribuidos[f"{c['mes']}_{c['ano']}"]),
                    'is_active': is_active_dict[f"{c['mes']}_{c['ano']}"],
                    'is_current': c['is_current']
                }
                for c in colunas_meses
            ]
        })

    # Preparar rodapé com totais por mês
    valores_totais_rodape = [
        _format_currency_br(totais_por_mes[f"{c['mes']}_{c['ano']}"]) 
        for c in colunas_meses
    ]

    anos_disponiveis = list(range(2024, 2033))
    
    total_global_acumulado_formatted = _format_currency_br(total_global_acumulado)

    context = {
        'dados_tabela':          dados_tabela,
        'colunas_meses':         colunas_meses,
        'valores_totais_rodape': valores_totais_rodape,
        'total_global_acumulado': total_global_acumulado_formatted,
        'total_meses_periodo':   total_meses_periodo,
        'mes_inicio':            mes_inicio,
        'ano_inicio':            ano_inicio,
        'mes_fim':               mes_fim,
        'ano_fim':               ano_fim,
        'meses_list':            [(i, MESES_PT[i]) for i in range(1, 13)],
        'anos_list':             anos_disponiveis,
    }

    # --- EXPORT HANDLERS ---
    export_format = request.GET.get('export')
    
    if export_format == 'csv':
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="distribuicao_mensal_{mes_inicio}_{ano_inicio}_a_{mes_fim}_{ano_fim}.csv"'
        response.write('\ufeff'.encode('utf8')) # UTF-8 BOM for Excel
        
        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        
        # Header Row
        headers = ['Contrato', 'Fornecedor', 'Secretaria', 'Valor Global']
        for col in colunas_meses:
            headers.append(col['label_compacto'])
        writer.writerow(headers)
        
        # Data Rows
        for row in dados_tabela:
            row_data = [row['numero'], row['fornecedor'], row['secretaria'], row['valor_global']]
            for val in row['valores_meses']:
                if val['is_active']:
                    row_data.append(val['formatted'])
                elif not row['tem_vigencia']:
                    row_data.append('Vigência não informada')
                else:
                    row_data.append('Fora da vigência')
            writer.writerow(row_data)
            
        # Footer Totals Row
        footer = ['TOTAL GERAL POR MÊS', '', '', total_global_acumulado_formatted]
        for total_mes_formatted in valores_totais_rodape:
            footer.append(total_mes_formatted)
        writer.writerow(footer)
        
        return response

    if export_format == 'excel':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Distribuição"

        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True

        # Styles
        font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color='003B70', end_color='003B70', fill_type='solid') # Navy primary
        
        font_data = Font(name='Arial', size=10)
        font_valor_global = Font(name='Arial', size=10, bold=True, color='003B70')
        font_zero = Font(name='Arial', size=9, color='94A3B8') # soft gray for out of vigencia
        
        font_footer = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        fill_footer = PatternFill(start_color='003B70', end_color='003B70', fill_type='solid') # institutional navy for totals
        
        border_thin = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        border_footer = Border(
            top=Side(style='medium', color='002244'),
            bottom=Side(style='medium', color='002244')
        )

        # Header Row
        headers = ['Contrato', 'Fornecedor', 'Secretaria', 'Valor Global']
        for col in colunas_meses:
            headers.append(col['label_compacto'])
            
        ws.append(headers)
        
        # Style Header
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center' if col_idx > 3 else 'left', vertical='center')
            cell.border = border_thin
        
        # Data Rows
        row_num = 2
        for row in dados_tabela:
            # Write Contract Metadata
            ws.cell(row=row_num, column=1, value=row['numero']).alignment = Alignment(horizontal='left')
            ws.cell(row=row_num, column=2, value=row['fornecedor']).alignment = Alignment(horizontal='left')
            ws.cell(row=row_num, column=3, value=row['secretaria']).alignment = Alignment(horizontal='left')
            
            # Valor Global (number formatted)
            cell_vg = ws.cell(row=row_num, column=4, value=float(row['valor_global_raw']))
            cell_vg.number_format = 'R$ #,##0.00'
            cell_vg.font = font_valor_global
            cell_vg.alignment = Alignment(horizontal='right')
            
            col_idx = 5
            for val in row['valores_meses']:
                cell_val = ws.cell(row=row_num, column=col_idx)
                if val['is_active']:
                    cell_val.value = float(val['raw'])
                    cell_val.number_format = 'R$ #,##0.00'
                    cell_val.font = font_data
                    cell_val.alignment = Alignment(horizontal='right')
                elif not row['tem_vigencia']:
                    cell_val.value = "Vigência não informada"
                    cell_val.font = font_zero
                    cell_val.alignment = Alignment(horizontal='center')
                else:
                    cell_val.value = "Fora da vigência"
                    cell_val.font = font_zero
                    cell_val.alignment = Alignment(horizontal='center')
                cell_val.border = border_thin
                col_idx += 1
                
            # Apply general border to data columns 1-4
            for c in range(1, 5):
                ws.cell(row=row_num, column=c).border = border_thin
                if c < 4:
                    ws.cell(row=row_num, column=c).font = font_data
            
            row_num += 1
            
        # Totalizer Footer Row
        ws.cell(row=row_num, column=1, value="TOTAL GERAL POR MÊS").alignment = Alignment(horizontal='left')
        
        # Merge first 3 columns of footer row
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        
        ws.cell(row=row_num, column=4, value=float(total_global_acumulado)).number_format = 'R$ #,##0.00'
        
        col_idx = 5
        for col in colunas_meses:
            col_key = f"{col['mes']}_{col['ano']}"
            ws.cell(row=row_num, column=col_idx, value=float(totais_por_mes[col_key])).number_format = 'R$ #,##0.00'
            col_idx += 1
            
        # Style Totalizer Footer Row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = font_footer
            cell.fill = fill_footer
            cell.border = border_footer
            if col_idx >= 4:
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
                
        # Auto-adjust columns width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # Find the max string length
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format == 'R$ #,##0.00' and isinstance(cell.value, (int, float)):
                    # Format approximation for width calculation
                    val_str = f"R$ {cell.value:,.2f}".replace(',', 'x').replace('.', ',').replace('x', '.')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Specific overrides for widths
        ws.column_dimensions['A'].width = 16 # Contrato
        ws.column_dimensions['B'].width = 30 # Fornecedor
        ws.column_dimensions['C'].width = 25 # Secretaria
        ws.column_dimensions['D'].width = 18 # Valor Global
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="distribuicao_mensal_{mes_inicio}_{ano_inicio}_a_{mes_fim}_{ano_fim}.xlsx"'
        wb.save(response)
        return response
    
    return render(request, 'contratos/controle_mensal/calculo.html', context)
